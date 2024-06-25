import re
from docxtpl import *
from docx.shared import Mm
import openpyxl
class parse():
    doc=DocxTemplate('template.docx')
    Vuln_info={}#漏洞信息
    hostnum=0 #存活主机
    vuln_nums=0 #漏洞发现数统计
    host_info={} #主机信息
    risk_host_num=0 #危险主机数量
    network_risk=0 #网络风险
    server_class=[]#服务分类
    app_class=[]#应用分类
    sys_class=[]#系统分类
    vuln_class=[]#漏洞分类
    time_class=[]#时间分类
    CVE_time_class=[]#cve年份分类
    system=[]#操作系统类别
    count=[]

    def __init__(self,input_filepath,output_filepath,item_name,report_time):
        self.input_filepath=input_filepath
        self.output_filepath=output_filepath
        self.item_name = item_name
        self.report_time=report_time
    def xlsx_gen(self):
        workbook = openpyxl.load_workbook(self.input_filepath)
        # 获取工作表
        messge_chart=workbook['综述信息']
        vuln_info_chart = workbook['漏洞信息']
        host_info_chart=workbook['主机信息']
        self.hostnum = messge_chart['B3'].value
        self.risk_host_num=messge_chart['D3'].value
        self.network_risk=messge_chart['C3'].value
        self.get_vuln_info(vuln_info_chart)
        self.get_host_info(host_info_chart)
        self.get_messge(messge_chart,6)




    def get_vuln_info(self,vuln_info_chart):
        ipv4_pattern = re.compile(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b')
        # 遍历工作表的每一行（从第三行开始）
        for row in vuln_info_chart.iter_rows(min_row=3, values_only=True):
            # 获取第二列的值作为键
            key = row[1]
            # 提取ipv4个数
            matches = ipv4_pattern.findall(row[10])
            host_num1 = f'{len(matches)}/{self.hostnum}'  # 主机受影响个数
            # 将当前行的其他列的值（从第二列开始）转换为一个列表
            value_list = list(row[2:])  #
            value_list[8] = host_num1
            a = f'{round(len(matches) / int(self.hostnum) * 100)}%'  # 获取影响主机百分比
            value_list.append(a)
            self.Vuln_info[key] = value_list
            # 获取漏洞出现统计
            self.vuln_nums = self.vuln_nums + value_list[9]
    def get_messge(self,messge_chart,min_row):
        block = []
        blocks=[]
        datas=[]
        for row in messge_chart.iter_rows(min_row=min_row, values_only=True, max_row=messge_chart.max_row):
            datas.append(list(row[1:]))
        for data in datas:
            # 如果当前行不是空行
            if not all(value is None for value in data):
                # 将当前行添加到当前块中
                block.append(data)
                # 如果当前行是空行，并且已经收集了行（即存在一个块）
            else:
                if block:
                    # 将当前块添加到blocks列表中
                    blocks.append(block[1:])
                    # 重置当前块
                    block = []
        if block:
            blocks.append(block[1:])
                    # 处理最后一个块（如果有的话）
        self.server_class=blocks[0]
        self.app_class=blocks[1]
        self.sys_class=blocks[2]
        self.vuln_class=blocks[3]
        self.time_class=blocks[4]
        self.CVE_time_class=blocks[5]
        self.system=blocks[6]

    def get_host_info(self,host_chart):
        for row in host_chart.iter_rows(min_row=3, values_only=True):
            # 获取第二列的值作为键
            key = row[1]
            value_list = list(row[2:])
            self.host_info[key] = value_list
    def write_report(self):
        doc = self.doc
        high_vuln_image = InlineImage(doc, r'image\high_vuln_image.png', width=Mm(4.2),height=Mm(4.2))
        medium_vuln_image = InlineImage(doc, r'image\medium_vuln_image.png', width=Mm(4.2), height=Mm(4.2))
        low_vuln_image = InlineImage(doc, r'image\low_vuln_image.png', width=Mm(4.2), height=Mm(4.2))
        host_high = InlineImage(doc, r'image\host_high.png', width=Mm(4.2), height=Mm(4.2))
        host_medium = InlineImage(doc, r'image\host_medium.png', width=Mm(4.2), height=Mm(4.2))
        host_low = InlineImage(doc, r'image\host_low.png', width=Mm(4.2), height=Mm(4.2))
        secure = InlineImage(doc, r'image\secure.png', width=Mm(4.2), height=Mm(4.2))
        if 8.0 <= self.network_risk <= 10.0:
            network = '非常危险'
            image=host_high
        elif 5.0 <= self.network_risk < 8.0:
            network = '比较危险'
            image=host_medium
        elif 1.0 <= self.network_risk < 5.0:
            network = '比较安全'
            image=host_low
        else:
            network = '非常安全'
            image=secure
        context = {
            'item_name': self.item_name,  #项目名称
            'host_num':self.hostnum,  #主机数量
            'Vulns': self.Vuln_info,
            'Hosts':self.host_info,
            'report_time':self.report_time,
            'high_vuln_image': high_vuln_image,   #高危漏洞图片
            'low_vuln_image':low_vuln_image,
            'medium_vuln_image':medium_vuln_image,
            'host_high':host_high,
            'host_medium':host_medium,
            'host_low':host_low,
            'secure':secure,
            'vuln_nums':self.vuln_nums,
            'network_risk':self.network_risk,
            'risk_host_num':self.risk_host_num,
            'network':network,
            'image':image,
            'system':self.system,
            'server_class': self.server_class,
            'app_class':self.app_class,
            'CVE_time_class':self.CVE_time_class,
            'time_class':self.time_class,
            'vuln_class':self.vuln_class,
            'sys_class':self.sys_class,
            'count':self.count
        }
        doc.render(context)
        doc.save(self.output_filepath)
        print('【+】 Save report success\n')


if __name__ == '__main__':
    input_filepath='index.xlsx'  #xlsx文件名
    output_filepath='漏洞扫描报告.docx'   #word报告输出文件名
    item_name='金鸡滩' #项目名称
    report_time='2023-10-10'#报告时间  格式2023-10-03
    apt_parse=parse(input_filepath,output_filepath,item_name,report_time)
    apt_parse.xlsx_gen()
    apt_parse.write_report()
